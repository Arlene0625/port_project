# -*- coding: utf-8 -*-
# Author : monster
import json
from openpyxl import load_workbook
from  Interface_Project_201812.common import File_Path
from  Interface_Project_201812.common.Forward_loan_request import Forward_Loan
from Interface_Project_201812.common.Load_Conf import Load_Conf
class Wb_get:
    def __init__(self):
        self.case_id=None
        self.title=None
        self.method=None
        self.url=None
        self.request_data = None
        self.expected = None
        self.actual=None
        self.result = None

class Read_Excel:
    def __init__(self,book_name):
        try:
            self.book_name=book_name
            self.workbook=load_workbook(book_name)
        except FileNotFoundError as e:
            print('{}未找到，请检查文件路径'.format(book_name))
            raise e
    def read_excel(self,sheet_name):
        # open_wb=load_workbook(self.book_name)#打开工作薄
        open_sheet=self.workbook[sheet_name]#打开具体sheet页
        excel_data=[]
        for item in range(2,open_sheet.max_row+1):
            case=Wb_get()
            case.case_id=open_sheet.cell(item,1).value
            case.title = open_sheet.cell(item, 2).value
            case.method = open_sheet.cell(item, 3).value
            case.url = open_sheet.cell(item, 4).value
            case.request_data = open_sheet.cell(item, 5).value
            case.expected = open_sheet.cell(item, 6).value
            excel_data.append(case)
        return excel_data
    def get_sheet(self):
        return self.workbook.sheetnames   #获取该工作薄的所有sheetname的列表

    def write_actual(self,sheet_name,case_id,actual,result):#根据sheet_name，case_id分别定位到sheet和行，然后给actual赋值并保存
        open_sheet = self.workbook[sheet_name]  # 打开具体sheet页
        for i in range(2,open_sheet.max_row+1):
            case_id_i=open_sheet.cell(i,1).value#获取case_id
            if case_id_i == case_id:  # 判断获取的case_id和传进来的case_id是否相等
                open_sheet.cell(i, 7).value = actual  # 写入传进来的actual的值到当前的sheet——name的actual列
                open_sheet.cell(i, 8).value = result  # 写入result的值
                self.workbook.save(filename=self.book_name)
                break

    def read_url(self,sheet_name):
        open_sheet = self.workbook[sheet_name]  # 打开具体sheet页
        for i in range(2,open_sheet.max_row+1):
            case = Wb_get()
            case.url = open_sheet.cell(i, 4).value

if __name__ == '__main__':
    get_excel= Read_Excel(File_Path.excel_path)
    pp=get_excel.get_sheet()#获取指定路径下的工作薄
    # print(pp)
    for sheet_name in pp:
        hh = get_excel.read_excel(sheet_name)
        # print(type(hh))
        print(sheet_name+'的测试用例的个数为：',len(hh))#统计并 打印单个sheetname下的用例个数
        for case in hh:
            print(type(case))
            print('case的信息：',case.__dict__)#打印case信息
            url = Load_Conf().load_conf('api','url_prefix')  # 获取配置文件的url
            Url = url + case.url  # 拼接配置文件的url和excel表的url
            # print(Url)
            request=Forward_Loan(Url=Url,Param=eval(case.request_data),Method=case.method)
            print('相应结果：',request.get_json())#调用get_json()函数
            # print('code的值：',type(request.get_json()['code'])) #返回的code类型为str
            # print('excepted的值',type(case.expected))#返回的expected类型为int
            print('相应正文：',request.get_text())
            if case.expected==eval(request.get_json()['code']):#eval转化code类型为int
                result='pass'
            else:
                result='fail'
            print(result)
            # get_excel.write_actual(sheet_name=sheet_name,case_id=case.case_id,actual=request.get_text(),result=result)



